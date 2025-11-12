from openpyxl import Workbook
import csv
import os
from fpdf import FPDF
from pywebio import *
from pywebio.input import *
from pywebio.output import *

grades_dict = {"AA": 10,"AB": 9,"BB": 8,"BC": 7,"CC": 6,"CD": 5,"DD": 4,"DD*": 4,"F": 0,"F*": 0,"I": 0,}    #a helper dictionary for the grades 
stud_dict = {}   # a dictionary that will store the student details in a mapped fashion 
courses_dict = {}    #a helper dictionary to find out the name of the courses corresponding to the respective course code
heading = ["Sub. Code", "Subject Name", "L-T-P", "CRD", "GRD"]

student_details = {'roll' : "1901EE55", 'name' : "Amelia Jonas" ,
            'year' : "2019", "programme" : "Bachelor of Technology" ,
            'course' : "Computer science and engineering"}

course_details = {
    "CS" : "Computer Science and Engineering" ,
    "EE" : "Electrical Engineering" ,
    "ME" : "Mechanical Engineering"
}


def cpi_calc(grades , credits):  # function to calculate the cpi upto a particular semester 
    tot_sum = 0.0
    cred_sum = 0
    for i in range(len(grades)):
        tot_sum += grades[i]*credits[i]
    for c in credits:
        cred_sum += c
    return round(tot_sum/cred_sum, 2)  #return the rounded cpi upto 2 decimal

def pre_computation():
    print("pre computation starting")
    with open("sample_input/subjects_master.csv", "r") as file:
        reader = csv.DictReader(file)
        for i in reader:
            courses_dict[i["subno"]] = {"subname": i["subname"], "ltp": i["ltp"], "crd": i["crd"]}    #setting the values in a courses dict for later use 

    with open("sample_input/names-roll.csv", "r") as file:
        reader = csv.DictReader(file)     #reading the file as a dictionary 
        for i in reader:
            stud_dict[i["Roll"]] = {"Name": i["Name"]}   # setting the values in a stud_dict for later use  

    with open("sample_input/grades.csv", "r") as file:
        reader = csv.DictReader(file)
        for i in reader:
            try:
                stud_dict[i["Roll"]][i["Sem"]]
            except KeyError:
                stud_dict[i["Roll"]][i["Sem"]] = {}
            stud_dict[i["Roll"]][i["Sem"]][i["SubCode"]] = {"Grade": i["Grade"].strip(), "Sub_Type": i["Sub_Type"]}  #setting the grade values in the student dictionary 
    print("Pre computation done")


def footer(pdf) :
    pdf.set_y(pdf.get_y() + 80)
    y = pdf.get_y()
    pdf.set_x(30)
    pdf.set_font('Arial', 'B', 16)
    pdf.text(pdf.get_x() , pdf.get_y(), "Date of issue: ")
    pdf.line(75, pdf.get_y(), 150, pdf.get_y())
    pdf.set_y(y - 10)
    pdf.set_x(700)
    pdf.cell(80, 1, "", 'B', 2, 'C')
    pdf.cell(80, 10, "Assistant Registrar (Academic) ", 0, 0, 'C')


def generate_pdf(pdf):
    pdf.add_page()
    start_index_x = 30
    pdf.set_font('Arial' , 'B' , 16)
    pdf.set_left_margin(20)
    pdf.set_right_margin(20)
    pdf.cell(0 , 700 , "" ,  1, 1)
    pdf.set_xy(20 , 10)
    pdf.cell(80, 80, "", 1,0,'C')
    pdf.set_xy(20 , 10)
    pdf.image('logo.png' ,pdf.get_x() + 5 , pdf.get_y() + 5 , 70 , 70 , "png", 'logo.png')
    pdf.set_xy(100 , 10)
    pdf.cell(630 , 80 , "" , 1 , 0 , 'C')
    pdf.set_xy(100 , 10)
    pdf.image('name.jpg', pdf.get_x() + 5, pdf.get_y() + 5 , 625, 70, "png", 'name.jpg' )
    pdf.set_xy(730 , 10)
    pdf.cell(80, 80, "", 1,0,'C')
    pdf.set_xy(735 , 10)
    pdf.image('logo.png' ,pdf.get_x() , pdf.get_y() + 5 , 70 , 70 , "png", 'logo.png')
    pdf.set_y(80)
    pdf.set_x(start_index_x)

def semester_name(pdf, name):
    pdf.set_font('Arial',"BU", 16)
    pdf.cell(30, 10, f"Semester {name}", 0, 2)
    pdf.set_font('Arial','',16)

def overall_credits_cell(pdf , details) :
    pdf.set_font('Arial', 'B', 16)
    pdf.cell(200 , 10, f"Credits Taken: {details['credits']}    Credits Cleared: {details['credits']}  SPI: {details['spi']}   CPI: {details['cpi']}", 1, 2)
    pdf.set_font('Arial' , '', 16)

def make_description(pdf, details):
    pdf.set_y(pdf.get_y())
    pdf.set_x(250)
    pdf.cell(400, 16,"" ,1, 0, 'C')
    pdf.set_x(280)
    pdf.set_font("Arial", 'B', 16)
    pdf.cell(120, 8, f"Roll No:  {details['roll']}" , 0, 0)
    pdf.cell(120, 8, f"Name:  {details['name']}" , 0, 0)
    pdf.cell(120, 8, f"Year of admission:  {details['year']}" , 0, 1)
    pdf.set_x(280)
    pdf.cell(120, 8, f"Programme:  {details['programme']}" , 0, 0)
    pdf.cell(120, 8, f"Course:  {details['course']}" , 0, 1)
    pdf.set_y(pdf.get_y() + 10)

def set_coordinates(pdf, x, y, sem):
    # print(int(sem/4) , (int(sem%3)-1))
    pdf.set_y(y + int((sem-1)/3)*150)
    pdf.set_x(x + (int((sem-1)%3))*240 + (int((sem-1)%3))*20)
    if (sem-1)%3 == 0 and sem!=1 : 
        make_line(pdf, pdf.get_y())
        # pdf.set_y(pdf.get_y() + 20)

def create_cell(pdf ,type, to, content):
    
    # print(type ,to , content)
    if type==1 :
        # print("this is type 1")
        pdf.cell(60, 10, str(content) , 1, to, 'C')
        return

    if type==2 :
        # print("this is type 2")
        pdf.cell(140, 10, str(content) , 1, to, 'C')
        return

    if type==3 :
        # print("this is type 3")
        pdf.cell(20, 10, str(content) , 1, to, 'C')
        return

    if type==4 :
        # print("this is type 4")
        pdf.cell(15, 10, str(content) , 1, to, 'C')
        return

    if type==5 :
        # print("this is type 5")
        pdf.cell(15, 10, str(content) , 1, to, 'C')
        return
    if type==6 :
        # print("this is type 6")
        pdf.cell(100 , 10, str(content), 1, to, 'C')


def create_table(start_x, pdf ,headers, table_body):
    # print(headers)
    # semester_name(pdf,1)
    pdf.set_x(start_x)
    pdf.set_font("Arial", 'BU', 16)
    i = 1
    for heading in headers :
            if i < 5:
                create_cell(pdf, i , 0, heading)
            else : 
                create_cell(pdf, i, 1 , heading)
            i = i+1
    pdf.set_x(start_x)
    pdf.set_font("Arial", '', 16)
    i = 1
    for row in table_body:
        for column in row :
            if i < 5:
                create_cell(pdf, i , 0, column)
            else : 
                create_cell(pdf, i, 1 , column)
            i = i+1
        pdf.set_x(start_x)
        i = 1

def make_line(pdf , y) :
    pdf.line(20, y, 810, y)

def generate_marksheet(start_roll , end_roll):
    pre_computation()
    if os.path.exists("output") == False:
        os.makedirs("output")
    prefix = start_roll[0:6]

    if(start_roll[4:6] != end_roll[4:6]) :
        put_warning("The students cannot be of different departments", closable=True)
        return 

    if end_roll[0:6] != prefix:
        put_warning("Please enter correct range of roll numbers", closable=True)
        return 

    start = int(start_roll[6:len(start_roll)])
    end = int(end_roll[6:len(end_roll)])
    if (start > end) :
        put_warning("Please input correct range of roll numbers" , closable=True)
        return 
    print(start , end)
    put_html("<h3>Your Marksheets are being generated...... </h3>")
    not_present_roll_no = []

    for i in range(start , end+1):
        curr_roll = prefix + str(int(i/10)) + str(int(i%10))
        print(curr_roll)
        if curr_roll not in stud_dict :
            not_present_roll_no.append(curr_roll)
            continue 
        pdf = FPDF('L' ,'mm' , (800 , 830))
        generate_pdf(pdf)
        pdf.set_y(pdf.get_y() + 20)
        student_details['roll'] = curr_roll
        student_details['course'] = course_details[curr_roll[4:6]]
        make_description(pdf , student_details)
        pdf.set_x(30)
        credits = [0,0,0,0,0,0,0,0]                #list to store the credits sum of a semester 
        total_credits = [0,0,0,0,0,0,0,0]          #list to store the credits sum till semester 
        spi = [0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0]    #list to store the spi of all the semesters 
        cpi = [0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0]    #list to store the cpi of all the semesters
        x_coordinate = pdf.get_x()
        y_coordinate = pdf.get_y()
        for j in range(1, 9):
            set_coordinates(pdf, x_coordinate , y_coordinate, j)
            creds = []
            grades = []
            l = 1
            try:
                stud_dict[curr_roll][str(j)]
            except KeyError:
                continue
            table_data = []
            for k in stud_dict[curr_roll][str(j)]: #k is the subject code 
                # ws.append([l,k,courses_dict[k]["subname"],courses_dict[k]["ltp"],courses_dict[k]["crd"],stud_dict[curr_roll][str(j)][k]["Sub_Type"],stud_dict[curr_roll][str(j)][k]["Grade"]])
                temp_data_row = []
                temp_data_row.append(k)
                temp_data_row.append(courses_dict[k]['subname'])
                temp_data_row.append(courses_dict[k]["ltp"])
                temp_data_row.append(courses_dict[k]["crd"])
                temp_data_row.append(grades_dict[stud_dict[curr_roll][str(j)][k]["Grade"]])
                table_data.append(temp_data_row)
                creds.append(int(courses_dict[k]["crd"]))
                grades.append(grades_dict[stud_dict[curr_roll][str(j)][k]["Grade"]])
                l += 1
            semester_name(pdf,j)
            create_table(pdf.get_x(), pdf, heading, table_data)
            for c in creds:
                credits[j-1]+=c
            spi[j-1]=cpi_calc(grades, creds)
            if j>1:
                total_credits[j-1]=total_credits[j-2]+credits[j-1]
                cpi[j-1]=cpi_calc(spi[:j],credits[:j])
            else:
                total_credits[j-1]=credits[j-1]
                cpi[j-1]=spi[j-1]
            details = {'credits' : credits[j-1] , 'spi' : spi[j-1], 'cpi' : cpi[j-1]}
            # pdf.set_y(pdf.get_y() + 5)
            overall_credits_cell(pdf, details)
        set_coordinates(pdf , x_coordinate , y_coordinate, 10)
        footer(pdf)
        pdf.output("transcriptsIITP/"+curr_roll+".pdf")
    print(not_present_roll_no)
    put_success("The required marksheets have been generated, please look in the output folder for the same" ,  closable= True)
    for roll in not_present_roll_no :
        put_info(f"Roll no {roll} was not found in the list" , closable=True)

def main() :
    req_action = "Generate Marksheets"
    while(req_action != "None") :

        req_action = actions(label = "Please select your action" ,
                    buttons=[{'label' : "Generate Marksheets",'value':"Generate Marksheets"} ,
                        {'label' : "I wish to exit",'value':"None"}
                    ])

        if (req_action == "Generate Marksheets"):
            start = input("Please enter the starting roll number", type = TEXT)
            end = input("Please enter the ending Roll numbetr to generate the report", TYPE=TEXT)
            generate_marksheet(start , end)
        else :
            put_success("Hope you liked the project")
            
start_server(main, port=3001)        
